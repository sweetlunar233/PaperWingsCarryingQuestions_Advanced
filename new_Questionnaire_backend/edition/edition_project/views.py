from django.shortcuts import render

# Create your views here.
from django.http import JsonResponse  
from django.views.decorators.http import require_http_methods  

import random
from django.shortcuts import render
from django.http import JsonResponse

# Create your views here.
from .models import BlankQuestion,ChoiceQuestion,ChoiceOption,RatingQuestion
from .models import Answer,BlankAnswer,ChoiceAnswer,RatingAnswer
from .models import Submission 

import json
from django.core.mail import BadHeaderError, send_mail
from django.http import HttpResponse, HttpResponseRedirect

from django.core.mail import EmailMessage

from itsdangerous import URLSafeTimedSerializer as utsr
import base64
from django.conf import settings as django_settings
from django.utils import timezone
from django.db import transaction 

from rest_framework.views import APIView
import itertools

from itertools import chain  
from operator import attrgetter 

from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import ChoiceQuestion
from rest_framework import status

import requests
from io import BytesIO
  
userServeAddress='http://81.70.184.96:7000'
managementServeAddress='http://81.70.184.96:7001'
editionServeAddress='http://81.70.184.96:7002'

@require_http_methods(["GET"])  
def health_check(request):  
    # 这里可以添加一些实际的健康检查逻辑  
    # 例如，检查数据库连接、缓存连接等  
    # 如果一切正常，返回200 OK  
    # 如果有错误，返回503 Service Unavailable或其他适当的状态码  
    return JsonResponse({"status": "UP"}, status=200)

#普通问卷的展示界面：
def display_answer_normal(request,username,questionnaireId,submissionId):
    # 调用 user 项目的 API 获取用户信息
    user_api_url = f'{userServeAddress}/user/{username}/'
    try:
        user_response = requests.get(user_api_url)
        user_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
        user_data = user_response.json()
    except requests.RequestException as e:
        return JsonResponse({'error': str(e)}, status=500)
    # user=User.objects.get(username=username)
    if user_data is None:
        return HttpResponse(content='User not found', status=404) 
    
    # 调用management项目的API获取问卷信息
    survey_api_url = f'{managementServeAddress}/survey/{questionnaireId}/'
    try:
        survey_response = requests.get(survey_api_url)
        survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
        survey_data = survey_response.json()
    except requests.RequestException as e:
        return JsonResponse({'error': str(e)}, status=500)
        
    # survey=Survey.objects.get(SurveyID=questionnaireId)
    if survey_data is None:
        return HttpResponse(content='Questionnaire not found', status=404)   
    
    submission=Submission.objects.get(SubmissionID=submissionId)
    if submission is None:
        return HttpResponse(content='Submission not found', status=404)  
    
    all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(Survey=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(Survey=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                    RatingQuestion.objects.filter(Survey=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())
                                                    
    # 将迭代器转换为列表 (按QuestionNumber递增排序)
    all_questions_list = list(all_questionList_iterator)
    all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

    #print(all_questions_list.length())
    questionList=[]
    #print(all_questions)
    for question in all_questions_list:
        if question["Category"]==1 or question["Category"]==2:    #选择题

            #该单选题的用户选项:当前问卷当前submission(如果用户未选，则找不到对应的答案记录)
            if question["Category"]==1:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])  #只有一条记录
                
                #用户未填该单选题
                if not optionAnswer_query.exists():
                    answer=-1
                #用户填了这个单选题，有一条答案记录
                else:
                    answer=optionAnswer_query.first().ChoiceOptions.OptionID
            
            #该多选题的用户选项:当前问卷当前submission
            else:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])#一或多条记录
                #用户未填该多选题
                if not optionAnswer_query.exists():answer=[]
                #用户填了这个多选题，有一条/多条答案记录
                else:
                    answer=[]
                    for optionAnswer in optionAnswer_query:
                        answer.append(optionAnswer.ChoiceOptions.OptionID)

            optionList=[]
            #所有选项
            options_query=ChoiceOption.objects.filter(Question=question["QuestionID"])
            for option in options_query:
                optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,'optionId':option.OptionID})
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                    'optionList':optionList,'Answer':answer})
            
        elif question["Category"]==3:                  #填空题
            #该填空题的用户答案:有且仅有一条记录
            blankAnswer_query=BlankAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该填空题
            if not blankAnswer_query.exists():
                answer=""
            else:
                answer=blankAnswer_query.first().Content
            
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],
                                    'correctAnswer':question["CorrectAnswer"],'Answer':answer})

        elif question["Category"]==4:                  #评分题
            #该评分题的用户答案:有且仅有一条记录
            ratingAnswer_query=RatingAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该评分题
            if not ratingAnswer_query.exists():
                answer=0
            else:
                #print("123")
                answer=ratingAnswer_query.first().Rate

            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'Answer':answer})

    data={'Title':survey_data.Title,'description':survey_data.Description,'questionList':questionList}
    return JsonResponse(data)


#考试问卷的展示界面：
def display_answer_test(request,username,questionnaireId,submissionId):
    print("start display_answer_test")
    # print(submissionId)


    # 调用 user 项目的 API 获取用户信息
    user_api_url = f'{userServeAddress}/user/{username}/'
    try:
        user_response = requests.get(user_api_url)
        user_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
        user_data = user_response.json()
    except requests.RequestException as e:
        return JsonResponse({'error': str(e)}, status=500)
    

    # user=User.objects.get(username=username)
    if user_data is None:
        return HttpResponse(content='User not found', status=404) 
    
    survey_api_url = f'{managementServeAddress}/survey/{questionnaireId}/'
    try:
        survey_response = requests.get(survey_api_url)
        survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
        survey_data = survey_response.json()
    except requests.RequestException as e:
        return JsonResponse({'error': str(e)}, status=500)
        
    # survey=Survey.objects.get(SurveyID=questionnaireId)
    if survey_data is None:
        return HttpResponse(content='Questionnaire not found', status=404)   
    
    submission=Submission.objects.get(SubmissionID=submissionId)
    if submission is None:
        return HttpResponse(content='Submission not found', status=404)  
    score=submission.Score
    
    all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                    RatingQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())
    
    # 将迭代器转换为列表 (按QuestionNumber递增排序)
    all_questions_list = list(all_questionList_iterator)
    all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

    questionList=[]
    #print(all_questions)
    for question in all_questions_list:
        if question["Category"]==1 or question["Category"]==2:    #选择题

            #该单选题的用户选项:当前问卷当前submission(如果用户未选，则找不到对应的答案记录)
            if question["Category"]==1:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])  #只有一条记录
                
                #用户未填该单选题
                if not optionAnswer_query.exists():
                    answer=-1
                #用户填了这个单选题，有一条答案记录
                else:
                    answer=optionAnswer_query.first().ChoiceOptions.OptionID
            
            #该多选题的用户选项:当前问卷当前submission
            else:
                optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])#一或多条记录
                #用户未填该多选题
                if not optionAnswer_query.exists():answer=[]
                #用户填了这个多选题，有一条/多条答案记录
                else:
                    answer=[]
                    for optionAnswer in optionAnswer_query:
                        answer.append(optionAnswer.ChoiceOptions.OptionID)

            optionList=[]
            #所有选项
            options_query=ChoiceOption.objects.filter(Question=question["QuestionID"])
            for option in options_query:
                optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,'optionId':option.OptionID})
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                    'optionList':optionList,'Answer':answer})
            
        elif question["Category"]==3:                  #填空题
            #该填空题的用户答案:有且仅有一条记录
            blankAnswer_query=BlankAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该填空题
            if not blankAnswer_query.exists():
                answer=""
            else:
                answer=blankAnswer_query.first().Content
            
            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],
                                    'correctAnswer':question["CorrectAnswer"],'Answer':answer})

        elif question["Category"]==4:                  #评分题
            #该评分题的用户答案:有且仅有一条记录
            ratingAnswer_query=RatingAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
            #用户未填该评分题
            if not ratingAnswer_query.exists():
                answer=0
            else:
                answer=ratingAnswer_query.first().Rate

            questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                    'isNecessary':question["IsRequired"],'score':question["Score"],'Answer':answer})


    data={'Title':survey_data.Title,'description':survey_data.Description,'questionList':questionList,'score':score}
    # print(questionList[0])
    return JsonResponse(data)


#问卷填写界面：向前端传输问卷当前暂存的填写记录
class GetStoreFillView(APIView):
    def get(self, request, *args, **kwargs):  
        # 从查询参数中获取userName和surveyID   
        userName = kwargs.get('userName')  
        surveyID = kwargs.get('surveyID')   
        submissionID=kwargs.get('submissionID')  

        # 调用 user 项目的 API 获取用户信息
        user_api_url = f'{userServeAddress}/user/{userName}/'
        try:
            user_response = requests.get(user_api_url)
            user_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
            user_data = user_response.json()
        except requests.RequestException as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        # user=User.objects.get(username=userName)
        if user_data is None:
            return HttpResponse(content='User not found', status=404) 
        
        survey_api_url = f'{managementServeAddress}/survey/{surveyID}/'
        try:
            survey_response = requests.get(survey_api_url)
            survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
            survey_data = survey_response.json()
        except requests.RequestException as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        # survey=Survey.objects.get(SurveyID=surveyID)
        if survey_data is None:
            return HttpResponse(content='Questionnaire not found', status=404) 
          
        
        #从问卷广场界面进入：查找该用户是否有该问卷未提交的填写记录
        if submissionID=="-1":
            submission_query=Submission.objects.filter(RespondentID=user_data.UserID,SurveyID=survey_data.SurveyID,Status='Unsubmitted')
            if submission_query.exists():
                submissionID=submission_query.first().SubmissionID  #找到未填写的记录
                duration=submission_query.first().Interval
                submission = submission_query.first()
                # newsubmissionID = submissionID
            
            else:      #不存在：创建一条新的填写记录
                submission=Submission.objects.create(SurveyID=survey_data.SurveyID,RespondentID=user_data.UserID,Status="Unsubmitted",
                                                    Interval=0)
                duration=0
                submissionID=submission.SubmissionID
                # newsubmissionID = submission.SubmissionID
                # return HttpResponse(content='Submission not existed', status=404) 
                #################huyanzhe

        #submissionID=-2时,只传回问卷题干(同问卷编辑的GET接口)
        elif submissionID=="-2":
            print("--here---")
            all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                    RatingQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionNumber','QuestionID').all())
                                                    
            # 将迭代器转换为列表  
            all_questions_list = list(all_questionList_iterator)
            all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

            questionList=[]

            for question in all_questions_list:
                if question["Category"]==1 or question["Category"]==2:    #选择题
                    optionList=[]
                    #将所有选项顺序排列
                    options_query=ChoiceOption.objects.filter(Question=question["QuestionID"]).order_by('OptionNumber')
                    for option in options_query:
                        optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,
                                       'optionID':option.OptionID,'MaxSelectablePeople':option.MaxSelectablePeople})
                    questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                     'optionList':optionList})
                
                elif question["Category"]==3:                  #填空题
                
                    questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'correctAnswer':question["CorrectAnswer"]})

                elif question["Category"]==4:                  #评分题
                    questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"]})

        
            data={'Title':survey_data.Title,'category':survey_data.Category,'TimeLimit':survey_data.TimeLimit,
                'description':survey_data.Description,'questionList':questionList}
            return JsonResponse(data)
        
        submission=Submission.objects.filter(SubmissionID=submissionID).first()
        print(submission.Interval)
        # print(submission)
        if not submission:
            return HttpResponse(content='Submission not found', status=404) 
    
        Title=survey_data.Title
        Description=survey_data.Description
        category=survey_data.Category
        TimeLimit=survey_data.TimeLimit
        #people=survey.QuotaLimit
        
        '''1.以下部分与问卷编辑界面的get函数类似，拿到题干'''
        '''2.拿到当前submissionID对应填写记录'''
        all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID','MaxSelectable').all(),
                                                    RatingQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())
                                                    
        all_questions_list = list(all_questionList_iterator)

        # 将迭代器转换为列表 (按QuestionNumber递增排序)--顺序展示
        if survey_data.IsOrder:
            all_questions_list.sort(key=lambda x: x['QuestionNumber']) 
        
        #print(all_questions_list.length())
        questionList=[]
        #print(all_questions)
        for question in all_questions_list:
            if question["Category"]==1 or question["Category"]==2:    #选择题
                #print(question['MaxSelectable'])
                print(question)
                print(question['OptionCnt'])

                #该单选题的用户选项:当前问卷当前submission(如果用户未选，则找不到对应的答案记录)
                if question["Category"]==1:
                    optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])  #只有一条记录
                    #用户未填该单选题
                    if not optionAnswer_query.exists():
                        answer=-1
                    #用户填了这个单选题，有一条答案记录
                    else:
                        answer=optionAnswer_query.first().ChoiceOptions.OptionID
                
                #该多选题的用户选项:当前问卷当前submission
                else:
                    optionAnswer_query=ChoiceAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])#一或多条记录
                    #用户未填该多选题
                    if not optionAnswer_query.exists():answer=[]
                    #用户填了这个多选题，有一条/多条答案记录
                    else:
                        answer=[]
                        for optionAnswer in optionAnswer_query:
                            answer.append(optionAnswer.ChoiceOptions.OptionID)

                optionList=[]
                #将所有选项顺序排列
                print("***")
                options_query=ChoiceOption.objects.filter(Question=question["QuestionID"]).order_by('OptionNumber')
                for option in options_query:
                    optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,
                                       'optionId':option.OptionID,'MaxSelectablePeople':option.MaxSelectablePeople})
                
                if survey_data.Category == 3 and survey_data.IsOrder == False: #选项乱序展示
                    random.shuffle(optionList)
                
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                     'optionList':optionList,'Answer':answer,'max':question['MaxSelectable']})
                
            elif question["Category"]==3:                  #填空题
                #该填空题的用户答案:有且仅有一条记录
                blankAnswer_query=BlankAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
                #用户未填该填空题
                if not blankAnswer_query.exists():
                    answer=""
                else:
                    answer=blankAnswer_query.first().Content
                
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],
                                     'correctAnswer':question["CorrectAnswer"],'Answer':answer})

            elif question["Category"]==4:                  #评分题
                
                #该评分题的用户答案:有且仅有一条记录
                ratingAnswer_query=RatingAnswer.objects.filter(Submission=submission,Question=question["QuestionID"])
                #用户未填该评分题
                if not ratingAnswer_query.exists():
                    answer=0
                else:
                    answer=ratingAnswer_query.first().Rate

                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'Answer':answer})
        
        #题干乱序展示
        if survey_data.Category == 3 and survey_data.IsOrder == False:
            random.shuffle(questionList)

        #传回题干和填写记录
        data={'Title':survey_data.Title,'category':survey_data.Category,'TimeLimit':survey_data.TimeLimit,
            'description':survey_data.Description,'questionList':questionList,'duration':submission.Interval, 'submissionID':submissionID}
        
        return JsonResponse(data)
        
from .serializers import SubmissionSerializer

#问卷填写界面：从前端接收用户的填写记录(POST)
def get_submission(request):
    if(request.method=='POST'):
        try:
            # print("start get_submission")
            body=json.loads(request.body)
            surveyID=body['surveyID']    #问卷id
            status=body['status']  #填写记录状态
            submissionID=body['submissionID']   #填写记录ID
            username=body['username']     #填写者
            submissionList=body['question']     #填写记录
            duration=body['duration']  
            score=body['score'] 

            # 新加的
            publishDate=body['date']  #日期

            survey_api_url = f'{managementServeAddress}/survey/{surveyID}/'
            try:
                survey_response = requests.get(survey_api_url)
                survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
                survey_data = survey_response.json()
            except requests.RequestException as e:
                return JsonResponse({'error': str(e)}, status=500)

            # survey=Survey.objects.get(SurveyID=surveyID)
            if survey_data is None:
                return HttpResponse(content='Questionnaire not found',status=404)
            
            # 调用 user 项目的 API 获取用户信息
            user_api_url = f'{userServeAddress}/user/{username}/'
            try:
                user_response = requests.get(user_api_url)
                user_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
                user_data = user_response.json()
            except requests.RequestException as e:
                return JsonResponse({'error': str(e)}, status=500)
            
            # user=User.objects.get(username=username)
            if user_data is None:
                return HttpResponse(content='User not found',status=404)
        

            #当前不存在该填写记录，创建：  //实际上用不到，在getStoreFill的时候就给不存在的submission创建新的Id了
            currentTime=timezone.now()
            if submissionID==-1:
                submission=Submission.objects.create(SurveyID=survey_data.SurveyID,RespondentID=user_data.UserID,
                                             SubmissionTime=currentTime,Status=status,
                                             Interval=duration,Score=score)
                print(submission.SubmissionTime)
            
            #已存在，删除填写记录的所有内容
            else:
                submission=Submission.objects.get(SubmissionID=submissionID)
                if submission is None:
                    return HttpResponse(content='Submission not found',status=404)
                submission.Score=score
                submission.Status=status
                submission.Interval=duration
                submission.SubmissionTime=currentTime    #更新为当前时间
                submission.save()
                ######################huyanzhe
                
                #所有选择题的填写记录
                ChoiceAnswer_query=ChoiceAnswer.objects.filter(Submission=submission)
                if ChoiceAnswer_query.exists():
                    for choiceAnswer in ChoiceAnswer_query:
                        choiceAnswer.delete()
                
                #所有填空题的填写记录
                BlankAnswer_query=BlankAnswer.objects.filter(Submission=submission)
                if BlankAnswer_query.exists():
                    for blankAnswer in BlankAnswer_query:
                        blankAnswer.delete()
                
                #所有评分题的填写记录
                RatingAnswer_query=RatingAnswer.objects.filter(Submission=submission)
                if RatingAnswer_query.exists():
                    for ratingAnswer in RatingAnswer_query:
                        ratingAnswer.delete()

            # 向management传输submission信息
            submission_save_url=f'{managementServeAddress}/survey/submission_save'
            data={'SurveyID':survey_data.SurveyID,'RespondentID':user_data.UserID,'SubmissionTime':currentTime,
                  'Status':status,'Score':score,'Interval':duration}
            
            try:
                response=request.post(submission_save_url,data)
                response.raise_for_status()
            except requests.exceptions.RequestException as e:
                print(f'Error deleting edition service:{e}')


            for submissionItem in submissionList:
                # print("TieZhu")
                questionID=submissionItem["questionID"]     #问题ID
                answer=submissionItem['value']        #用户填写的答案
                category=submissionItem['category']     #问题类型（用于后续区分，解决不同种类问题的QuestionID会重复的问题）

                #print(category)
                #question = BaseQuestion.objects.get(QuestionID=questionID).select_subclasses()   #联合查询

                '''
                question_iterator=itertools.chain(ChoiceQuestion.objects.filter(QuestionID=questionID),
                                                    BlankQuestion.objects.filter(QuestionID=questionID),
                                                    RatingQuestion.objects.filter(QuestionID=questionID))
                question_list=list(question_iterator)
                question=question_list[0]
                print(question)
                print(question_list)
                # print(question["Category"])
                # print(question.Category)'''

                questionNewList=[]
                choiceQuestion_query=ChoiceQuestion.objects.filter(QuestionID=questionID,Category=category)
                if choiceQuestion_query.exists():
                    questionNewList.append(choiceQuestion_query.first())

                blankQuestion_query=BlankQuestion.objects.filter(QuestionID=questionID,Category=category)
                if blankQuestion_query.exists():
                    questionNewList.append(blankQuestion_query.first())

                ratingQuestion_query=RatingQuestion.objects.filter(QuestionID=questionID,Category=category)
                if ratingQuestion_query.exists():
                    questionNewList.append(ratingQuestion_query.first())
                
                question=questionNewList[0]
                
                # print("123154654")

                # print(question.CorrectAnswer)
                if question is None:
                    return HttpResponse(content='Question not found',status=404)

                if question.Category==1:     #单选题：Answer为选项ID
                    if answer==-1: continue       #返回-1，代表用户没填该单选题
                    option=ChoiceOption.objects.get(OptionID=answer)     #用户选择的选项
                    if option is None:
                        return HttpResponse(content="Option not found",status=404)
                    choiceAnswer=ChoiceAnswer.objects.create(Question=question,Submission=submission,ChoiceOptions=option)
                    choiceAnswer.save()

                    #若已提交，报名问卷的必填选择题中，选择的对应选项人数-1
                    if status=='Submitted' and survey_data.Category==2 and question.IsRequired==True:
                        print(option.MaxSelectablePeople)

                        if option.MaxSelectablePeople<=0:
                            data={'message':False,'content':'报名人数已满'}
                            return JsonResponse(data)
                        
                        else:
                            option.MaxSelectablePeople-=1
                            option.save()


                elif question.Category==2:     #多选题：Answer为选项ID的数组
                    #为每个用户选择的选项，创建一条ChoiceAnswer记录
                    for optionID in answer:
                        option=ChoiceOption.objects.get(OptionID=optionID)     #用户选择的选项
                        if option is None:
                            return HttpResponse(content="Option not found",status=404)
                        choiceAnswer=ChoiceAnswer.objects.create(Question=question,Submission=submission,ChoiceOptions=option)
                        choiceAnswer.save()

                        #若已提交，报名问卷的必填选择题中，选择的对应选项人数-1
                        if status=='Submitted' and survey_data.Category==2 and question.IsRequired==True:
                            if option.MaxSelectablePeople<=0:
                                data={'message':False,'content':'报名人数已满'}
                                return JsonResponse(data)
                            else:
                                option.MaxSelectablePeople-=1
                                option.save()

                elif question.Category==3:     #填空题：answer为填写的内容
                    blankAnswer=BlankAnswer.objects.create(Question=question,Submission=submission,Content=answer)
                    blankAnswer.save()
                
                elif question.Category==4:      #评分题：answer为填写的内容
                    # print(answer)
                    ratingAnswer=RatingAnswer.objects.create(Question=question,Submission=submission,Rate=answer)
                    ratingAnswer.save()

            user_data.zhibi+=50
            user_data.save()

            # 向user传输更新后的zhibi信息
            user_save_url=f'{userServeAddress}/user_save/'
            data={'UserID':user_data.UserID,'zhibi':user_data.zhibi}
            
            try:
                response=request.post(user_save_url,data)
                response.raise_for_status()
            except requests.exceptions.RequestException as e:
                print(f'Error deleting edition service:{e}')
            ######################huyanzhe
                
        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    data={'message':True,'submissionId':submissionID}
    # print(submissionID)
    return JsonResponse(data)
    #return JsonResponse({'error': 'Invalid request method'}, status=405)


#问卷编辑界面：向前端传输问卷设计内容
class GetQuestionnaireView(APIView):
    def get(self, request, survey_id, *args, **kwargs):  
        design = request.GET.get('design', 'false')  # 默认为'false'  
        design = design.lower() == 'true'  # 将字符串转换为布尔值  

        survey_api_url = f'{managementServeAddress}/survey/{survey_id}/'
        try:
            survey_response = requests.get(survey_api_url)
            survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
            survey_data = survey_response.json()
        except requests.RequestException as e:
            return JsonResponse({'error': str(e)}, status=500)
        # survey=Survey.objects.get(SurveyID=survey_id)
        if survey_data is None:
            return HttpResponse(content='Questionnaire not found', status=400) 
        title=survey_data.Title
        catecory=survey_data.Category
        #people=survey.QuotaLimit
        TimeLimit=survey_data.TimeLimit

        '''
        blank_questions = list(BlankQuestion.objects.filter(Survey=survey).values_list('id', 'QuestionNumber'))  
        choice_questions = list(ChoiceQuestion.objects.filter(Survey=survey).values_list('id', 'QuestionNumber'))  
        rating_questions = list(RatingQuestion.objects.filter(Survey=survey).values_list('id', 'QuestionNumber'))  

        # 将这些列表合并，并基于QuestionNumber进行排序  
        combined_questions = sorted(chain(blank_questions, choice_questions, rating_questions), key=lambda x: x[1])
        '''

        all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','MaxSelectable','QuestionID').all(),
                                                    RatingQuestion.objects.filter(SurveyID=survey_data.SurveyID).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionNumber','QuestionID').all())
                                                    
        # 将迭代器转换为列表  
        all_questions_list = list(all_questionList_iterator)
        all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

        questionList=[]

        #print(all_questions)
        for question in all_questions_list:
            if question["Category"]==1 or question["Category"]==2:    #选择题
                optionList=[]
                #将所有选项顺序排列
                options_query=ChoiceOption.objects.filter(Question=question["QuestionID"]).order_by('OptionNumber')
                for option in options_query:
                    optionList.append({'content':option.Text,'optionNumber':option.OptionNumber,'isCorrect':option.IsCorrect,
                                       'optionID':option.OptionID,'MaxSelectablePeople':option.MaxSelectablePeople})
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'optionCnt':question["OptionCnt"],
                                     'optionList':optionList,'max':question['MaxSelectable']})
                
            elif question["Category"]==3:                  #填空题
                
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"],'correctAnswer':question["CorrectAnswer"]})

            elif question["Category"]==4:                  #评分题
                questionList.append({'type':question["Category"],'question':question["Text"],'questionID':question["QuestionID"],
                                     'isNecessary':question["IsRequired"],'score':question["Score"]})

        
        data={'Title':survey_data.Title,'category':survey_data.Category,'TimeLimit':survey_data.TimeLimit,
              'description':survey_data.Description,'questionList':questionList}
        
        return JsonResponse(data, status=200)


#问卷编辑界面：从前端接收问卷的设计内容
def save_qs_design(request):
    if(request.method=='POST'):
        try:
            body=json.loads(request.body)
            surveyID=body['surveyID']    #问卷id
            title=body['title']  #问卷标题
            catecory=body['category']   #问卷类型（普通0、投票1、报名2、考试3）
            isOrder=body['isOrder'] #是否顺序展示（考试问卷）
            #people=body['people']   #报名人数（报名问卷）
            timelimit=body['timeLimit']
            username=body['userName']   #创建者用户名
            description=body['description'] #问卷描述
            Is_released=body['Is_released'] #保存/发布

            questionList=body['questionList']   #问卷题目列表

            # 新加的
            publishDate=body['date'] #日期

            print(questionList)

            # 调用 user 项目的 API 获取用户信息
            user_api_url = f'{userServeAddress}/user/{username}/'
            try:
                user_response = requests.get(user_api_url)
                user_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
                user_data = user_response.json()
            except requests.RequestException as e:
                return JsonResponse({'error': str(e)}, status=500)
            # user=User.objects.get(username=username)
            if user_data is None:        
                return HttpResponse(content='User not found', status=400) 
            
            #当前不存在该问卷，创建：
            if surveyID==-1:
                # survey=Survey.objects.create(OwnerID=user_data.UserID,Title=title,
                #                              Description=description,Is_released=Is_released,
                #                              Is_open=True,Is_deleted=False,Category=catecory,
                #                              TotalScore=0,TimeLimit=timelimit,IsOrder=isOrder
                #                             )
                print("TieZhu")
                #survey.QuotaLimit=people
                ###############huyanzhe
                url = f'{managementServeAddress}/survey/update-survey'
                data = {
                    'SurveyID': surveyID,
                    'OwnerID': user_data.UserID,
                    'Title': title,
                    'Description': description,
                    'Is_released': Is_released,
                    'Is_open': True,
                    'Is_deleted': False,
                    'Category': catecory,
                    'TotalScore': 0,
                    'TimeLimit': timelimit,
                    'IsOrder': isOrder,
                    'PublishDate': None
                }
                try:
                    response = requests.post(url, data)
                    response.raise_for_status()
                    response_data = response.json()
                    id = response_data['SurveyID']
                    print(f"Successfully deleted edition service: {response.json()}")
                except requests.exceptions.RequestException as e:
                    print(f"Error deleting edition service: {e}")
            #已有该问卷的编辑记录
            else:
                #####################huyanzhe
                url = f'{managementServeAddress}/survey/update-survey'
                data = {
                    'SurveyID': surveyID,
                    'OwnerID': None,
                    'Title': title,
                    'Description': description,
                    'Is_released': Is_released,
                    'Is_open': None,
                    'Is_deleted': None,
                    'Category': catecory,
                    'TotalScore': None,
                    'TimeLimit': timelimit,
                    'IsOrder': isOrder,
                    'PublishDate': publishDate
                }
                try:
                    response = requests.post(url, data)
                    response.raise_for_status()
                    response_data = response.json()
                    id = response_data['SurveyID']
                    print(f"Successfully deleted edition service: {response.json()}")
                except requests.exceptions.RequestException as e:
                    print(f"Error deleting edition service: {e}")

                #该问卷的所有选择题
                choiceQuestion_query=ChoiceQuestion.objects.filter(SurveyID=id)
                for choiceQuestion in choiceQuestion_query:
                    #删除该选择题的所有选项
                    choiceOption_query=ChoiceOption.objects.filter(Question=choiceQuestion)
                    for choiceOption in choiceOption_query:
                        choiceOption.delete()
                    choiceQuestion.delete()

                #删除该问卷的所有填空题
                blankQuestion_query=BlankQuestion.objects.filter(SurveyID=id)
                for blankQuestion in blankQuestion_query:
                    blankQuestion.delete()
                
                #删除该问卷的所有评分题
                ratingQuestion_query=RatingQuestion.objects.filter(SurveyID=id)
                for ratingQuestion in ratingQuestion_query:
                    ratingQuestion.delete()

            index=1
            for question in questionList:
                if question["type"]==1 or question["type"]==2:        #单选/多选

                    optionList=question['optionList']
                    
                    question=ChoiceQuestion.objects.create(SurveyID=id,Text=question["question"],IsRequired=question["isNecessary"],
                                                                QuestionNumber=index,Score=question["score"],Category=question["type"],
                                                                OptionCnt=question["optionCnt"],MaxSelectable=question['max'])
                    question.save()

                    #所有选项:
                    jdex=1
                    for option in optionList:
                        option=ChoiceOption.objects.create(Question=question,Text=option["content"],IsCorrect=option["isCorrect"],
                                                           OptionNumber=jdex,MaxSelectablePeople=option['MaxSelectablePeople'])
                        option.save()
                        jdex=jdex+1

                
                elif question["type"]==3:                          #填空
                    # print(question)
                    question=BlankQuestion.objects.create(SurveyID=id,Text=question["question"],IsRequired=question["isNecessary"],
                                                        Score=question["score"],QuestionNumber=index,Category=question["type"],
                                                            CorrectAnswer=question["correctAnswer"])
                    question.save()  
                
                else:                                           #评分题
                    question=RatingQuestion.objects.create(SurveyID=id,Text=question["question"],IsRequired=question["isNecessary"],
                                                              Score=question["score"],QuestionNumber=index,Category=question["type"])
                    question.save()
                index=index+1
            return HttpResponse(content='Questionnaire saved successfully', status=200) 
        except json.JSONDecodeError:  
            return JsonResponse({'error': 'Invalid JSON body'}, status=400)
        except Exception as e:  
            return JsonResponse({'error': str(e)}, status=500) 
    return JsonResponse({'error': 'Invalid request method'}, status=405)
    
#交叉分析
def cross_analysis(request, QuestionID1, QuestionID2):
    if request.method == 'GET':
        
        question1 = ChoiceQuestion.objects.filter(QuestionID=QuestionID1).first()
        question2 = ChoiceQuestion.objects.filter(QuestionID=QuestionID2).first()

        if QuestionID1 is None or QuestionID2 is None:
            return JsonResponse({'error': 'Missing QuestionID(s)'}, status=404)
        
        if question1.Survey.SurveyID != question2.Survey.SurveyID:
            return JsonResponse({'error': 'Two questions are not from the same questionnaire.'}, status=404)
        
        if question1.Category!=1 and question1.Category!=2:
            return JsonResponse({'error':'Question1 is not a choice question.'},status=404)
        
        if question2.Category!=1 and question2.Category!=2:
            return JsonResponse({'error':'Question2 is not a choice question.'},status=404)
        

        #问题1的所有选项
        all_options_iterator_1=itertools.chain(ChoiceOption.objects.filter(Question=question1).values('OptionID','OptionNumber','Text').all())
                                  
        all_options_list_1 = list(all_options_iterator_1)
        all_options_list_1.sort(key=lambda x: x['OptionNumber']) 

        #问题2的所有选项
        all_options_iterator_2=itertools.chain(ChoiceOption.objects.filter(Question=question2).values('OptionID','OptionNumber','Text').all())
                                  
        all_options_list_2 = list(all_options_iterator_2)
        all_options_list_2.sort(key=lambda x: x['OptionNumber']) 


        crossCount=[]      #数组：选每个交叉选项的人数
        crossText=[]       #数组：每个交叉选项的内容
        for option1 in all_options_list_1:
            for option2 in all_options_list_2:
                crossText.append('-'.join([option1['Text'],option2['Text']]))
                # crossCount.append('-'.join([str(ChoiceAnswer.objects.filter(Question=question1,ChoiceOptions=option1['OptionID']).count()),
                #                             str(ChoiceAnswer.objects.filter(Question=question2,ChoiceOptions=option2['OptionID']).count())]))
                submission_option1 = ChoiceAnswer.objects.filter(Question=question1,ChoiceOptions=option1['OptionID']).values_list('Submission', flat=True).distinct()
                cnt = ChoiceAnswer.objects.filter(Submission__in=submission_option1,Question=question2,ChoiceOptions=option2['OptionID']).count()
                crossCount.append(cnt)
        data={'crossCount':crossCount,'crossText':crossText}
        
        return JsonResponse(data)

#下载表格
from openpyxl import Workbook
import datetime
from urllib.parse import quote  

def download_submissions(request, surveyID):
    if request.method == 'GET':
        try:
            survey_api_url = f'{managementServeAddress}/survey/{surveyID}/'
            try:
                survey_response = requests.get(survey_api_url)
                survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
                survey_data = survey_response.json()
            except requests.RequestException as e:
                return JsonResponse({'error': str(e)}, status=500)
            # survey = Survey.objects.get(SurveyID=surveyID)
            if survey_data is None:
                return HttpResponse(content='Questionniare not found.', status=404)  

            wb = Workbook()
            wb.encoding='utf-8'
            ws=wb.active    #获取第一个工作表


            #第一行内容
            content=['填写者','提交答卷时间','所用时间']
            ws.cell(1,1).value='填写者'
            ws.cell(1,2).value='提交答卷时间'
            ws.cell(1,3).value='所用时间'

            #该问卷的所有问题
            all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(SurveyID=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                        ChoiceQuestion.objects.filter(SurveyID=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID').all(),
                                                        RatingQuestion.objects.filter(SurveyID=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionNumber','QuestionID').all())
                                                    
            # 将迭代器转换为列表  
            all_questions_list = list(all_questionList_iterator)
            all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

            for i in range(0,len(all_questions_list)):
                question=all_questions_list[i]
                ws.cell(1,i+4,'、'.join([str(i+1),question['Text']])).value='、'.join([str(i+1),question['Text']])    #题号+题干

            #找出该问卷的所有已提交/已打分的填写记录，按提交时间从前向后排序
            submission_list=list(Submission.objects.filter(SurveyID=survey_data, Status__in=['Submitted', 'Graded']).values(
                'SubmissionID','SubmissionTime','Respondent','Survey','Interval','Score'
            ))
            if len(submission_list)==0:
                return HttpResponse(content='No submission records available.',status=404)
        
            submission_list.sort(key=lambda x: x['SubmissionTime'])   

            sub_index=1
            ques_index=3
            for submission in submission_list:
                sub_index+=1
                # 调用 user 项目的 API 获取用户信息
                id = submission.RespondentID
                user_api_url = f'{userServeAddress}/user/{id}/'
                try:
                    user_response = requests.get(user_api_url)
                    user_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
                    user_data = user_response.json()
                except requests.RequestException as e:
                    return JsonResponse({'error': str(e)}, status=500)
                # user=User.objects.get(UserID=submission['Respondent'])
                ws.cell(sub_index,1).value=user_data.username

                ws.cell(sub_index,2).value=submissionTime=submission['SubmissionTime'].strftime('%Y-%m-%d %H:%M:%S')
            
                if survey_data.Category==3:  #考试问卷：所用时间为Duration；其他问卷：'--'
                    ws.cell(sub_index,3).value=submission['Interval']
                else:
                    ws.cell(sub_index,3).value='--'
            
                ques_index=3
                for question in all_questions_list:
                    ques_index+=1

                    if question['Category']==1 or question['Category']==2:
                        #该题的所有选项
                        all_options_answer_list=list(ChoiceAnswer.objects.filter(Question=question['QuestionID'],Submission=submission['SubmissionID'])
                                                      .values('ChoiceOptions'))

                        if len(all_options_answer_list)==0:   #没有选择选项
                            ws.cell(sub_index,ques_index).value=' '

                        else: 
                            optionNumberList=[]
                            for answer in all_options_answer_list:
                                option=ChoiceOption.objects.filter(OptionID=answer['ChoiceOptions']).first()
                                optionNumberList.append(str(option.OptionNumber))
                            optionNumberList.sort()     #选项顺序升序排列

                            ws.cell(sub_index,ques_index).value=','.join(optionNumberList)
                    
                    elif question['Category']==3:      #填空题
                        answer=BlankAnswer.objects.get(Question=question['QuestionID'],Submission=submission['SubmissionID'])
                        if answer is None:   
                            ws.cell(sub_index,ques_index).value=' '
                        else: 
                            ws.cell(sub_index,ques_index).value=answer.Content

                    else:           #评分题
                        answer=RatingAnswer.objects.get(Question=question['QuestionID'],Submission=submission['SubmissionID'])
                        if answer is None:   
                            ws.cell(sub_index,ques_index).value=' '
                        else: 
                            ws.cell(sub_index,ques_index).value=answer.Rate

                if survey_data.Category==3:      #考试问卷：增加该填写记录的总得分
                    ws.cell(sub_index,ques_index).value=submission['Score']

            # 准备写入到IO中
            output = BytesIO()
            #ws.save()
            wb.save(output)     #Excel文件内容保存到IO中

            output.seek(0)	 # 重新定位到开始

            # 设置HttpResponse的类型
            response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
            formatted_now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            file_name = '-'.join([survey_data.Title, '填写记录统计']) + '_' + formatted_now + '.xlsx' 

            file_name = quote(file_name)	 # 使用urlquote()方法解决中文无法使用的问题
            response['Content-Disposition'] = 'attachment; filename=%s' % file_name
            return response

        except Exception as e:
            return JsonResponse({'message': '导出表格失败!'})

    return JsonResponse({'error': 'Invalid request method'}, status=405)

from django.db.models import Count, Sum, Q

#数据分析：向前端传输该问卷的所有题目及填写内容的统计数据
def survey_statistics(request, surveyID):
    if (request.method=='GET'):
        #问卷
        survey_api_url = f'{managementServeAddress}/survey/{surveyID}/'
        try:
            survey_response = requests.get(survey_api_url)
            survey_response.raise_for_status()  # 如果请求失败，将引发 HTTPError
            survey_data = survey_response.json()
        except requests.RequestException as e:
            return JsonResponse({'error': str(e)}, status=500)
        # survey = Survey.objects.filter(SurveyID=surveyID)
        # if not survey.exists():
        #     return JsonResponse({'error': 'Survey not found'}, status=404)
        if survey_data is None:
            return HttpResponse(content='Questionniare not found.', status=404)
        
        # survey = Survey.objects.filter(SurveyID=surveyID).first()
        # print("lorian")
        
        
        all_questionList_iterator = itertools.chain(BlankQuestion.objects.filter(SurveyID=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','CorrectAnswer','QuestionNumber','QuestionID').all(),
                                                    ChoiceQuestion.objects.filter(SurveyID=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','OptionCnt','QuestionNumber','QuestionID','MaxSelectable').all(),
                                                    RatingQuestion.objects.filter(SurveyID=survey_data).values('Category', 'Text', 'QuestionID', 'IsRequired', 'Score','QuestionID','QuestionNumber').all())

        # 将迭代器转换为列表 (按QuestionNumber递增排序)                                        
        all_questions_list = list(all_questionList_iterator)
        all_questions_list.sort(key=lambda x: x['QuestionNumber']) 

        #print(all_questions_list.length())
        questionList=[]
        #print(all_questions)
        for question in all_questions_list:
            if question["Category"]==1 or question["Category"]==2:    #选择题
                # print("选择题")
                # print(question)
                 
                #该问题的所有选项
                # all_options_iterator=itertools.chain(ChoiceOption.objects.filter(Question=question).values('OptionNumber','Text'))               
                # all_options_list = list(all_options_iterator)

                all_options_list = list(ChoiceOption.objects.filter(Question=question['QuestionID']).values('OptionID','OptionNumber','Text'))
                all_options_list.sort(key=lambda x: x['OptionNumber']) 

                optionCount=[]      #数组：选每个选项的人数
                optionText=[]       #数组：每个选项的内容
                for option in all_options_list:
                    optionText.append(option['Text'])
                    optionCount.append(ChoiceAnswer.objects.filter(Question=question['QuestionID'],ChoiceOptions=option['OptionID']).count())

                questionList.append({'Content':question["Text"],'Text':optionText,'Count':optionCount,'type':question['Category'],"questionId":question['QuestionID']})

                
                    
            elif question["Category"]==3:       #填空题
                answerText=[]   #数组：填空的内容
                answerCount=[]  #数组：填该内容的人数


                #text_list=list(BlankAnswer.objects.filter(Question=question['QuestionID']).values('Content'))

                text_counts = BlankAnswer.objects.filter(Question=question['QuestionID']).values('Content').annotate(count=Count('Content'))
                for item in text_counts:
                    answerText.append(item['Content'])
                    answerCount.append(item['count'])

                questionList.append({'Content':question["Text"],'Text':answerText,'Count':answerCount,'type':question['Category'],"questionId":question['QuestionID']})

            
            else:       #评分题
                answerText=[]   #数组：评分的分数
                answerCount=[]  #数组：该评分的人数

                rate_counts=RatingAnswer.objects.filter(Question=question['QuestionID']).values('Rate').annotate(count=Count('Rate'))
                for item in rate_counts:
                    answerText.append(item['Rate'])
                    answerCount.append(item['count'])

                questionList.append({'Content':question["Text"],'Text':answerText,'Count':answerCount,'type':question['Category'],"questionId":question['QuestionID']})


        #传回题干和填写记录的统计数据
        data={'title':survey_data.Title,'category':survey_data.Category,'TimeLimit':survey_data.TimeLimit,
            'description':survey_data.Description,'questionList':questionList}
        return JsonResponse(data)

    
@api_view(['POST'])
def UpdateSubmissionStatus(request, submission_id):
    if not submission_id:
        return Response({'error': 'SubmissionID and Status are required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        submission = Submission.objects.get(SubmissionID=submission_id)
        submission.Status = 'Deleted'
        submission.save()
        return Response({'status': 'Submission updated'}, status=status.HTTP_200_OK)
    except Submission.DoesNotExist:
        return Response({'error': 'Submission not found'}, status=status.HTTP_404_NOT_FOUND)
    

@api_view(['POST'])
def check_survey_status(request):
    survey_id = request.data.get('survey_id')
    if not survey_id:
        return Response({'error': 'survey_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        choice_questions = ChoiceQuestion.objects.filter(Survey_id=survey_id, Category__in=[1,2], IsRequired=True)
        is_full = True

        for choice_question in choice_questions:
            choice_options = ChoiceOption.objects.filter(Question=choice_question)
            for choice_option in choice_options:
                if choice_option.MaxSelectablePeople > 0:
                    is_full = False
                    break
            if not is_full:
                break

        return Response({'is_full': is_full})

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def DeleteSubmission(request, SubmissionID):
    try:
        submission = Submission.objects.get(SubmissionID=SubmissionID)
        submission.delete()
        submission.save()
    except Submission.DoesNotExist:
        return Response({'error': 'Survey not found'}, status=status.HTTP_404_NOT_FOUND)

    return JsonResponse({'message': "success"})