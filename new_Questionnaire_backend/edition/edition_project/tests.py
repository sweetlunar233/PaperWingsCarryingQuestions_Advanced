from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from .models import Submission,BlankQuestion,ChoiceQuestion,RatingQuestion,ChoiceOption
from .models import ChoiceAnswer,RatingAnswer,BlankAnswer
import json
from rest_framework.test import APIClient
from io import BytesIO
import openpyxl
import datetime

# Create your tests here.

class save_qs_design_test(TestCase):
    def setUp(self):
        self.url = reverse('save-qs-design-url')
        self.questionList = [
            {"type": 1, "question": "Choice question", "isNecessary": True, "score": 10, "optionCnt": 3, "max": 2, "optionList": [
                {"content": "Option 1", "isCorrect": True, "MaxSelectablePeople": 50},
                {"content": "Option 2", "isCorrect": False, "MaxSelectablePeople": 50}
            ]},
            {"type": 3, "question": "Blank question", "isNecessary": True, "score": 5, "correctAnswer": 1}
        ]

        self.survey_data = {
            "surveyID": -1, "title": "New Survey", "category": 0, "isOrder": True, "timeLimit": 30, "userName": "lorian",
            "description": "Test survey", "Is_released": False, "questionList": self.questionList
        }

    def test_create_new_survey(self):
        # 测试创建新问卷
        response = self.client.post(self.url, json.dumps(self.survey_data), content_type='application/json')
        self.assertEqual(response.status_code, 200)

    def test_invalid_request_method(self):
        # 测试错误的请求方法
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method')

class get_questionnaire_test(TestCase):
    def setUp(self):
        # 创建测试环境
        self.client = APIClient()
        self.url = reverse('get_questionnaire', args=[1])

    def test_get_questionnaire_success(self):
        # 测试成功获取问卷设计内容
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)

# -----问卷填写界面------

class get_store_fill_test(TestCase):
    def setUp(self):
        # 创建测试环境
        self.client = APIClient()
        self.url =reverse('get-store-fill-url', kwargs={'userName': 'lorian', 'surveyID': 1, 'submissionID': -1})

    def test_get_existing_submission(self):
        # 测试获取现有的问卷填写记录
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        response_data = response.json()  # 获取完整的响应数据
        self.assertIn('Title', response_data)

    def test_questionnaire_not_found(self):
        # 测试问卷不存在的情况
        url = reverse('get-store-fill-url', kwargs={'userName': 'lorian', 'surveyID': 9999, 'submissionID': 1})  # 不存在的 SurveyID
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)

class get_submission_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
        self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
        self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)
        self.choice_option = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1, MaxSelectablePeople=10)
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Unsubmitted')

    def test_successful_submission(self):
        # 测试成功提交填写记录
        submission_data = {
            'surveyID': self.survey.SurveyID,
            'status': 'Submitted',
            'submissionID': 1,
            'username': self.user.username,
            'question': [
                {'questionID': self.choice_question.QuestionID, 'value': self.choice_option.OptionID, 'category': 1},
                {'questionID': self.blank_question.QuestionID, 'value': 'Answer', 'category': 3},
                {'questionID': self.rating_question.QuestionID, 'value': 4, 'category': 4},
            ],
            'duration': 120,
            'score': 0,
            'date': timezone.now().isoformat()
        }

        # 使用json.dumps将数据序列化为JSON字符串
        response = self.client.post(reverse('post-submission-url'), data=json.dumps(submission_data), content_type='application/json')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get('message'), True)

        # 检查选项选择记录是否保存
        self.assertTrue(ChoiceAnswer.objects.filter(Question=self.choice_question, Submission__SubmissionID=1).exists())
        # 检查填空题答案是否保存
        self.assertTrue(BlankAnswer.objects.filter(Question=self.blank_question, Submission__SubmissionID=1).exists())
        # 检查评分题答案是否保存
        self.assertTrue(RatingAnswer.objects.filter(Question=self.rating_question, Submission__SubmissionID=1).exists())
        # 检查用户纸币是否增加
        self.user.refresh_from_db()
        self.assertEqual(self.user.zhibi, 150)

    def test_invalid_json_body(self):
        # 测试无效的 JSON 数据
        response = self.client.post(reverse('post-submission-url'), data="Invalid JSON", content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get('error'), 'Invalid JSON body')

# #问卷展示界面：

# class display_answer_normal_test(TestCase):
#     def setUp(self):
#         # 设置测试环境
#         self.client = APIClient()
#         self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
#         self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
#         self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Submitted')

#         # 创建问题和答案
#         self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
#         self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
#         self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)
        
#         self.choice_option = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
#         self.choice_answer = ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option)
#         self.blank_answer = BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Test Answer")
#         self.rating_answer = RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

#         # 设置URL
#         self.url = reverse('display-answer-normal', kwargs={'username': self.user.username, 'questionnaireId': self.survey.SurveyID, 'submissionId': self.submission.SubmissionID})

#     def test_display_answer_normal_success(self):
#         # 测试成功展示问卷答案
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, 200)
#         self.assertIn('Choice Question', response.json()['questionList'][0]['question'])
#         self.assertIn('Blank Question', response.json()['questionList'][1]['question'])
#         self.assertIn('Rating Question', response.json()['questionList'][2]['question'])

#     def test_user_not_found(self):
#         # 测试用户不存在的情况
#         url = reverse('display-answer-normal', kwargs={'username': 'unknownuser', 'questionnaireId': self.survey.SurveyID, 'submissionId': self.submission.SubmissionID})
#         response = self.client.get(url)
#         self.assertEqual(response.status_code, 404)
#         self.assertIn('User not found', response.content.decode())

# class display_answer_test_test(TestCase):
#     def setUp(self):
#         # 设置测试环境
#         self.client = APIClient()
#         self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
#         self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
#         self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Submitted', Score=80)

#         # 创建问题和答案
#         self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
#         self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
#         self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)
        
#         self.choice_option = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
#         self.choice_answer = ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option)
#         self.blank_answer = BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Test Answer")
#         self.rating_answer = RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

#         # 设置URL
#         self.url = reverse('display-answer-test', kwargs={'username': self.user.username, 'questionnaireId': self.survey.SurveyID, 'submissionId': self.submission.SubmissionID})

#     def test_display_answer_test_success(self):
#         # 测试成功展示考试问卷答案
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, 200)
#         response_data = response.json()

#         self.assertIn('Choice Question', response_data['questionList'][0]['question'])
#         self.assertIn('Blank Question', response_data['questionList'][1]['question'])
#         self.assertIn('Rating Question', response_data['questionList'][2]['question'])
#         self.assertEqual(response_data['score'], 80)

#     def test_survey_not_found(self):
#         # 测试问卷不存在的情况
#         url = reverse('display-answer-test', kwargs={'username': self.user.username, 'questionnaireId': 9999, 'submissionId': self.submission.SubmissionID})
#         response = self.client.get(url)
#         self.assertEqual(response.status_code, 404)
#         self.assertIn('Questionnaire not found', response.content.decode())

# #数据分析:

# class cross_analysis_test(TestCase):
#     def setUp(self):
#         # 设置测试环境
#         self.client = APIClient()
#         self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
#         self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, Is_released=True)

#         # 创建问题
#         self.question1 = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question 1", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
#         self.question2 = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question 2", Score=5, IsRequired=True, QuestionNumber=2, Category=1)

#         # 创建选项
#         self.option1_q1 = ChoiceOption.objects.create(Question=self.question1, Text="Option 1.1", IsCorrect=True, OptionNumber=1)
#         self.option2_q1 = ChoiceOption.objects.create(Question=self.question1, Text="Option 1.2", IsCorrect=False, OptionNumber=2)

#         self.option1_q2 = ChoiceOption.objects.create(Question=self.question2, Text="Option 2.1", IsCorrect=True, OptionNumber=1)
#         self.option2_q2 = ChoiceOption.objects.create(Question=self.question2, Text="Option 2.2", IsCorrect=False, OptionNumber=2)

#         # 创建提交记录和答案
#         self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Submitted', Score=80)
#         ChoiceAnswer.objects.create(Question=self.question1, Submission=self.submission, ChoiceOptions=self.option1_q1)
#         ChoiceAnswer.objects.create(Question=self.question2, Submission=self.submission, ChoiceOptions=self.option1_q2)

#         # 设置URL
#         self.url = reverse('cross-analysis-url', kwargs={'QuestionID1': self.question1.QuestionID, 'QuestionID2': self.question2.QuestionID})

#     def test_cross_analysis_success(self):
#         # 测试交叉分析的成功情况
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, 200)
#         response_data = response.json()

#         self.assertIn('crossCount', response_data)
#         self.assertIn('crossText', response_data)
#         self.assertEqual(len(response_data['crossCount']), 4)
#         self.assertEqual(len(response_data['crossText']), 4)

#     def test_questions_from_different_surveys(self):
#         # 测试两个问题来自不同问卷的情况
#         survey2 = Survey.objects.create(Owner=self.user, Title="Another Survey", Category=1, TimeLimit=30, Is_released=True)
#         question3 = ChoiceQuestion.objects.create(Survey=survey2, Text="Choice Question 3", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
#         url = reverse('cross-analysis-url', kwargs={'QuestionID1': self.question1.QuestionID, 'QuestionID2': question3.QuestionID})
#         response = self.client.get(url)
#         self.assertEqual(response.status_code, 404)
#         self.assertIn('Two questions are not from the same questionnaire.', response.content.decode())

# class survey_statistics_test(TestCase):
#     def setUp(self):
#         # 设置测试环境
#         self.client = APIClient()
#         self.user = User.objects.create(username='testuser', email='user@example.com', password='password123')
#         self.survey = Survey.objects.create(Owner=self.user, Title="Survey Statistics", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)

#         # 创建问题
#         self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
#         self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
#         self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)

#         # 创建选项
#         self.choice_option_1 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
#         self.choice_option_2 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 2", IsCorrect=False, OptionNumber=2)

#         # 创建提交记录和答案
#         self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Submitted', Score=80)
#         ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option_1)
#         BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Blank Answer")
#         RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

#         # 设置URL
#         self.url = reverse('survey-statistics-url', kwargs={'surveyID': self.survey.SurveyID})

#     def test_survey_statistics_success(self):
#         # 测试获取问卷统计数据的成功情况
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, 200)
#         response_data = response.json()

#         self.assertIn('title', response_data)
#         self.assertEqual(response_data['title'], self.survey.Title)
#         self.assertIn('questionList', response_data)
#         self.assertEqual(len(response_data['questionList']), 3)

#         # 检查选择题统计
#         choice_question_data = next(item for item in response_data['questionList'] if item['type'] == 1)
#         self.assertIn('Content', choice_question_data)
#         self.assertEqual(choice_question_data['Content'], self.choice_question.Text)
#         self.assertEqual(len(choice_question_data['Text']), 2)
#         self.assertEqual(len(choice_question_data['Count']), 2)

#         # 检查填空题统计
#         blank_question_data = next(item for item in response_data['questionList'] if item['type'] == 3)
#         self.assertIn('Content', blank_question_data)
#         self.assertEqual(blank_question_data['Content'], self.blank_question.Text)
#         self.assertEqual(blank_question_data['Text'][0], "Blank Answer")

#         # 检查评分题统计
#         rating_question_data = next(item for item in response_data['questionList'] if item['type'] == 4)
#         self.assertIn('Content', rating_question_data)
#         self.assertEqual(rating_question_data['Content'], self.rating_question.Text)
#         self.assertEqual(rating_question_data['Text'][0], 4)

#     def test_survey_not_found(self):
#         # 测试问卷不存在的情况
#         url = reverse('survey-statistics-url', kwargs={'surveyID': 999})
#         response = self.client.get(url)
#         self.assertEqual(response.status_code, 404)
#         self.assertIn('Survey not found', response.content.decode())

# class download_submissions_test(TestCase):
#     def setUp(self):
#         # 设置测试环境
#         self.client = APIClient()
#         self.user = User.objects.create(username='testuser', email='user@example.com', password='password123')
#         self.survey = Survey.objects.create(Owner=self.user, Title="Survey Statistics", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)

#         # 创建问题
#         self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
#         self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
#         self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)

#         # 创建选项
#         self.choice_option_1 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
#         self.choice_option_2 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 2", IsCorrect=False, OptionNumber=2)

#         # 创建提交记录和答案
#         self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Submitted', Interval=0, Score=80)
#         ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option_1)
#         BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Blank Answer")
#         RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

#         # 设置URL
#         self.url = reverse('download_submissions-url', kwargs={'surveyID': self.survey.SurveyID})

#     def test_download_submissions_success(self):
#         # 测试成功下载提交记录
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, 200)
#         self.assertEqual(response['Content-Type'], 'application/vnd.ms-excel')

#         # 读取返回的Excel文件
#         output = BytesIO(response.content)
#         workbook = openpyxl.load_workbook(output)
#         worksheet = workbook.active

#         # 检查Excel中的内容
#         self.assertEqual(worksheet.cell(1, 1).value, '填写者')
#         self.assertEqual(worksheet.cell(1, 2).value, '提交答卷时间')
#         self.assertEqual(worksheet.cell(1, 3).value, '所用时间')
#         self.assertEqual(worksheet.cell(2, 1).value, self.user.username)
#         self.assertEqual(worksheet.cell(2, 3).value, '--') # 对于非考试问卷，其所用时间为'--'
#         self.assertEqual(worksheet.cell(2, 4).value, '1')  # 选项号1
#         self.assertEqual(worksheet.cell(2, 5).value, 'Blank Answer')
#         self.assertEqual(worksheet.cell(2, 6).value, 4)

#     def test_no_submissions(self):
#         # 测试无提交记录的情况
#         Submission.objects.all().delete()
#         response = self.client.get(self.url)
#         self.assertEqual(response.status_code, 404)
#         self.assertIn('No submission records available.', response.content.decode())
