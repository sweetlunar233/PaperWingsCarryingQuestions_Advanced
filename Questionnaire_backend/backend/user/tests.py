from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from .models import User,Survey,Submission,RewardOffering,BlankQuestion,ChoiceQuestion,RatingQuestion,ChoiceOption
from .models import ChoiceAnswer,RatingAnswer,BlankAnswer
import json
from rest_framework.test import APIClient
from io import BytesIO
import openpyxl
import datetime

class send_registration_email_test(TestCase):
    def setUp(self):
        # 设置测试数据
        # 正数据
        self.valid_payload = {
            'username': 'test',
            'password': 'test',
            'email': '1378832571@qq.com'
        }

        # 创建一个已存在的用户
        User.objects.create(username='TieZhu', email='dcx1378832571@163.com', password='test')

        # 反数据
        self.existing_username = {
            'username': 'TieZhu', # 已存在的用户名
            'password': 'test',
            'email': 'dcx1378832571@163.com'
        }



    def test_register_new_user(self):
        # 测试有效用户注册
        url = reverse('send-registration-email-url') 
        response = self.client.post(url, json.dumps(self.valid_payload), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['message'], "register success")

    def test_register_with_existing_username(self):
        # 测试已存在用户名注册
        url = reverse('send-registration-email-url')
        response = self.client.post(url, json.dumps(self.existing_username), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['message'], "same username")

class get_user_info_test(TestCase):
    def setUp(self):
        # 创建测试用户
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)

    def test_get_user_info_success(self):
        # 成功获取用户信息
        url = reverse('get-user-info-url', kwargs={'username': 'testuser'})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            str(response.content, encoding='utf8'),
            {
                'password': 'password123',
                'email': 'user@example.com',
                'zhibi': 100,
                'photo': self.user.get_used_element()
            }
        )

    def test_invalid_method(self):
        # 测试错误的请求方法
        url = reverse('get-user-info-url', kwargs={'username': 'testuser'})
        response = self.client.post(url)  # 使用POST方法尝试
        self.assertEqual(response.status_code, 405)
        self.assertJSONEqual(
            str(response.content, encoding='utf8'),
            {'error': 'Invalid request method'}
        )

class modify_user_info_test(TestCase):
    def setUp(self):
        # 创建一个已存在的用户
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)

        # 正数据，用于修改用户信息
        self.valid_payload = {
            'username': 'testuser',
            'flag': 1,
            'email': 'user@example.com',
            'password': 'password123'
        }

        # 反数据

        # 错误的请求数据
        self.invalid_payload = {
            'username': 'testuser',
            'flag': -1, #参数不正确
        }

    def test_modify_user_info_success(self):
        # 测试成功修改用户信息
        url = reverse('modify-user-info-url')
        response = self.client.post(url, json.dumps(self.valid_payload), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['message'], "True")

    def test_invalid_parameters(self):
        # 测试错误的参数
        url = reverse('modify-user-info-url')
        response = self.client.post(url, json.dumps(self.invalid_payload), content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Invalid or missing parameters')

class modify_photo_in_shop_test(TestCase):
    def setUp(self):
        # 创建测试用户
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100, own_photos=json.dumps(['photo1', 'photo2']))
        self.url = reverse('modify-photo-in-shop-url')

        # 正确的测试数据
        self.valid_payload = {
            'username': 'testuser',
            'photonumber': 1,
            'status': 'newstatus',
            'money': 500
        }

    def test_modify_user_success(self):
        # 测试成功修改用户头像和纸币信息
        response = self.client.post(self.url, json.dumps(self.valid_payload), content_type='application/json')
        self.user.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.user.zhibi, 500)
        self.assertIn('newstatus', json.loads(self.user.own_photos))

    def test_invalid_json_body(self):
        # 测试提交的 JSON 数据无效
        response = self.client.post(self.url, '{"username": "testuser", "photonumber": "}', content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Invalid JSON body')

    def test_invalid_request_method(self):
        # 测试使用非法的请求方法
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method')

# -----------------------

class get_drafted_qs_test(TestCase):
    def setUp(self):
        # 创建测试用户
        self.user=User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey1 = Survey.objects.create(Owner=self.user, Is_released=False, Title="Survey 1", PublishDate=timezone.now(), SurveyID=1, Category=1)
        self.survey2 = Survey.objects.create(Owner=self.user, Is_released=True, Title="Survey 2", PublishDate=timezone.now(), SurveyID=2, Category=2)
        self.url = reverse('get-drafted-qs-url', kwargs={'username': 'testuser'})

    def test_get_draft_surveys(self):
        # 测试获取草稿调查
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()['data']), 1)  # 只有一个草稿
        self.assertEqual(response.json()['data'][0]['Title'], 'Survey 1')

    def test_invalid_request_method(self):
        # 测试错误的请求方法
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method')

class get_released_qs_test(TestCase):
    def setUp(self):
        # 创建测试用户和问卷
        self.user = User.objects.create(username='testuser')
        self.survey1 = Survey.objects.create(Owner=self.user, Is_released=True, Is_deleted=False, Title="Survey 1",
                                             PublishDate=timezone.now(), SurveyID=1, Category=1,
                                             Description="A test survey", Is_open=True)
        self.survey2 = Survey.objects.create(Owner=self.user, Is_released=True, Is_deleted=True, Title="Survey 2",  # 已删除的问卷
                                             PublishDate=timezone.now(), SurveyID=2, Category=2,
                                             Description="A deleted survey", Is_open=False)
        Submission.objects.create(Survey=self.survey1,Respondent=self.user,
                                             SubmissionTime=timezone.now(),Status='Submitted',
                                             Interval=2,Score=5)

        self.url = reverse('get-released-qs-url', kwargs={'username': 'testuser'})

    def test_get_released_surveys(self):
        # 测试获取已发布的问卷
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()['data']), 1)  # 只有一个未删除的问卷
        self.assertEqual(response.json()['data'][0]['FilledPeople'], 1)  # 验证填写人数统计正确

    def test_invalid_request_method(self):
        # 测试错误的请求方法
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method')

class get_filled_qs_test(TestCase):
    def setUp(self):
        # 创建测试用户和问卷及提交记录
        self.user = User.objects.create(username='testuser')
        self.survey = Survey.objects.create(Owner=self.user, Title="Survey 1",Is_released=True,
                                            PublishDate=timezone.now(), SurveyID=1, Category=1,
                                            Description="A test survey", Is_open=True)
        Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionTime=timezone.now(),
                                  Status='Submitted', SubmissionID=1)
        self.url = reverse('get-filled-qs-url', kwargs={'username': 'testuser'})

    def test_get_filled_surveys(self):
        # 测试获取用户填写的问卷
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        data = response.json()['data']
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['Status'], '已提交')
        self.assertEqual(data[0]['SubmissionID'], 1)

    def test_invalid_request_method(self):
        # 测试错误的请求方法
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method') 

class get_all_released_qs_test(TestCase):
    def setUp(self):
        # 创建测试用户和问卷及可能的奖励
        self.user = User.objects.create(username='testuser')
        self.survey1 = Survey.objects.create(Owner=self.user, Title="Open Survey 1",
                                             PublishDate=timezone.now(), SurveyID=1, Category=1,
                                             Description="Open for everyone", Is_released=True, Is_open=True)
        self.survey2 = Survey.objects.create(Owner=self.user, Title="Closed Survey 2",
                                             PublishDate=timezone.now(), SurveyID=2, Category=2,
                                             Description="Not open", Is_released=True, Is_open=False)
        RewardOffering.objects.create(Survey=self.survey1, Zhibi=100, AvailableQuota=50)
        self.url = reverse('get-all-released-qs-url')

    def test_get_all_released_surveys(self):
        # 测试获取所有已发布的开放问卷
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        data = response.json()['data']
        self.assertEqual(len(data), 1)  # 只有一个问卷是已发布且开放的
        self.assertEqual(data[0]['SurveyID'], 1)
        self.assertEqual(data[0]['Reward'], 100)  # 验证奖励信息正确

    def test_invalid_request_method(self):
        # 测试错误的请求方法
        response = self.client.post(self.url, {})
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method')

class check_qs_test(TestCase):
    def setUp(self):
        # 创建测试用户和问卷
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Survey", Category=1,Is_released=True,
                                            PublishDate=timezone.now(), SurveyID=1, Is_open=True)
        self.url = reverse('check-qs-url', kwargs={'username': 'testuser', 'questionnaireId': 1,'type':1})

    def test_submission_exists_and_unsubmitted(self):
        # 测试已存在未提交的提交记录
        Submission.objects.create(Respondent=self.user, Survey=self.survey, Status='Unsubmitted')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content.decode(), {'message': False, "content": "对于当前问卷，您有未提交的填写记录"})

    def test_submission_exists_and_submitted(self):
        # 测试已存在已提交的提交记录
        Submission.objects.create(Respondent=self.user, Survey=self.survey, Status='Submitted')
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content.decode(), {'message': False, "content": "您完成投票，不可重复投票"})

    def test_can_fill_survey(self):
        # 测试用户可以开始或继续填写问卷
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content.decode(), {'message': "True", "content": "可以开始/继续填写"})

# -----------------------

class delete_unreleased_qs_test(TestCase):
    def setUp(self):
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        # 创建未发布的问卷
        self.survey = Survey.objects.create(Owner=self.user,SurveyID=1, Title="Unreleased Survey", Is_released=False)
        self.url = reverse('delete-unreleased-qs-url')

    def test_delete_unreleased_survey(self):
        # 测试成功删除未发布的问卷
        response = self.client.post(self.url, json.dumps(1), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['message'], "True")
        self.assertFalse(Survey.objects.filter(SurveyID=1).exists())  # 确认问卷已被删除

    def test_no_id_provided(self):
        # 测试没有提供ID的情况
        response = self.client.post(self.url, json.dumps(None), content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'No ID provided')

    def test_questionnaire_not_found(self):
        # 测试找不到指定问卷
        response = self.client.post(self.url, json.dumps(999), content_type='application/json')
        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()['error'], 'No questionnaire found with the given ID')

class update_or_delete_released_qs(TestCase):
    def setUp(self):
        # 创建测试用户和问卷
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, SurveyID=1, Is_released=True, Is_open=True, Is_deleted=False)
        self.url = reverse('delete-released-qs-url')

        # 创建相关的提交记录
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Unsubmitted')

    def test_delete_released_survey(self):
        # 测试删除已发布的问卷
        data = {'flag': 1, 'id': 1}
        response = self.client.post(self.url, json.dumps(data), content_type='application/json')
        self.survey.refresh_from_db()
        self.submission.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.survey.Is_deleted)
        self.assertFalse(self.survey.Is_released)
        self.assertEqual(self.submission.Status, 'Deleted')

    def test_update_released_survey_status(self):
        # 测试更新问卷发布状态
        data = {'flag': 2, 'id': 1}
        response = self.client.post(self.url, json.dumps(data), content_type='application/json')
        self.survey.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.survey.Is_open)  # 撤回发布状态

    def test_invalid_json_body(self):
        # 测试提交的 JSON 数据无效
        response = self.client.post(self.url, '{"bad_json": "}', content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Invalid JSON body')

class delete_filled_qs_test(TestCase):
    def setUp(self):
        # 创建测试用户和问卷及提交记录
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Survey", Is_released=True)
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1)
        self.url = reverse('delete-filled-qs')

    def test_delete_submission(self):
        # 测试成功删除提交记录
        data = 1
        response = self.client.post(self.url, json.dumps(data), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['message'], "True")
        self.assertFalse(Submission.objects.filter(SubmissionID=1).exists())  # 确认提交记录已被删除

    def test_invalid_json_body(self):
        # 测试提交的 JSON 数据无效
        response = self.client.post(self.url, '{"bad_json": "}', content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['error'], 'Invalid JSON body')

# -----------------------

class save_qs_design_test(TestCase):
    def setUp(self):
        # 创建测试用户
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.url = reverse('save-qs-design-url')
        self.questionList = [
            {"type": 1, "question": "Choice question", "isNecessary": True, "score": 10, "optionCnt": 3, "max": 2, "optionList": [
                {"content": "Option 1", "isCorrect": True, "MaxSelectablePeople": 50},
                {"content": "Option 2", "isCorrect": False, "MaxSelectablePeople": 50}
            ]},
            {"type": 3, "question": "Blank question", "isNecessary": True, "score": 5, "correctAnswer": 1}
        ]
        
        # 设置时区感知的日期
        date_str = '2021-01-01'
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        aware_datetime = timezone.make_aware(datetime.datetime.combine(date_obj, datetime.time(0, 0)))

        self.survey_data = {
            "surveyID": -1, "title": "New Survey", "category": 0, "isOrder": True, "timeLimit": 30, "userName": "testuser",
            "description": "Test survey", "Is_released": False, "questionList": self.questionList, "date": aware_datetime.isoformat()
        }

    def test_create_new_survey(self):
        # 测试创建新问卷
        response = self.client.post(self.url, json.dumps(self.survey_data), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(Survey.objects.filter(Title="New Survey").exists())

    def test_update_existing_survey(self):
        # 测试更新已存在的问卷
        existing_survey = Survey.objects.create(Owner=self.user, Title="Old Survey", Is_released=True, Category=0)
        self.survey_data['surveyID'] = existing_survey.SurveyID
        response = self.client.post(self.url, json.dumps(self.survey_data), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        updated_survey = Survey.objects.get(SurveyID=existing_survey.SurveyID)
        self.assertEqual(updated_survey.Title, "New Survey")

    def test_invalid_request_method(self):
        # 测试错误的请求方法
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 405)
        self.assertEqual(response.json()['error'], 'Invalid request method')

class get_questionnaire_test(TestCase):
    def setUp(self):
        # 创建测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123')
        self.survey = Survey.objects.create(Owner=self.user, Is_released=True, Title="Survey", Category=1, TimeLimit=30, SurveyID=1)
        self.url = reverse('get_questionnaire', args=[self.survey.SurveyID])
        
        # 添加问题，确保为所有相关字段提供值
        BlankQuestion.objects.create(Survey=self.survey, Text="What is your name?", Category=3, Score=10, IsRequired=True, QuestionNumber=1)
        choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choose one:", Category=1, Score=5, IsRequired=False, QuestionNumber=2, OptionCnt=1)
        ChoiceOption.objects.create(Question=choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
        RatingQuestion.objects.create(Survey=self.survey, Text="Rate our service:", Category=4, Score=5, IsRequired=True, QuestionNumber=3)

    def test_get_questionnaire_success(self):
        # 测试成功获取问卷设计内容
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "What is your name?")
        self.assertContains(response, "Choose one:")
        self.assertContains(response, "Rate our service:")

    def test_request_with_design_parameter(self):
        # 测试设计参数
        url = f"{self.url}?design=true"
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Option 1")

#问卷填写界面：

class get_store_fill_test(TestCase):
    def setUp(self):
        # 创建测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123')
        self.survey = Survey.objects.create(Owner=self.user, Title="Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
        self.url = lambda submission_id: reverse('get-store-fill-url', kwargs={'userName': self.user.username, 'surveyID': 1, 'submissionID': submission_id})

        # 创建一个默认的问卷填写记录
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Unsubmitted')

    def test_get_existing_submission(self):
        # 测试获取现有的问卷填写记录
        response = self.client.get(self.url(-2))
        self.assertEqual(response.status_code, 200)

    def test_questionnaire_not_found(self):
        # 测试问卷不存在的情况
        url = reverse('get-store-fill-url', kwargs={'userName': self.user.username, 'surveyID': 999, 'submissionID': 1})  # 不存在的 SurveyID
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)
        self.assertIn('Questionnaire not found', response.content.decode())

class get_submission_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
        self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
        self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)
        self.choice_option = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1, MaxSelectablePeople=10)
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Unsubmitted')

    def test_successful_submission(self):
        # 测试成功提交填写记录
        submission_data = {
            'surveyID': self.survey.SurveyID,
            'status': 'Submitted',
            'submissionID': 1,
            'username': self.user.username,
            'question': [
                {'questionID': self.choice_question.QuestionID, 'value': self.choice_option.OptionID, 'category': 1},
                {'questionID': self.blank_question.QuestionID, 'value': 'Answer', 'category': 3},
                {'questionID': self.rating_question.QuestionID, 'value': 4, 'category': 4},
            ],
            'duration': 120,
            'score': 0,
            'date': timezone.now().isoformat()
        }

        # 使用json.dumps将数据序列化为JSON字符串
        response = self.client.post(reverse('post-submission-url'), data=json.dumps(submission_data), content_type='application/json')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json().get('message'), True)

        # 检查选项选择记录是否保存
        self.assertTrue(ChoiceAnswer.objects.filter(Question=self.choice_question, Submission__SubmissionID=1).exists())
        # 检查填空题答案是否保存
        self.assertTrue(BlankAnswer.objects.filter(Question=self.blank_question, Submission__SubmissionID=1).exists())
        # 检查评分题答案是否保存
        self.assertTrue(RatingAnswer.objects.filter(Question=self.rating_question, Submission__SubmissionID=1).exists())
        # 检查用户纸币是否增加
        self.user.refresh_from_db()
        self.assertEqual(self.user.zhibi, 150)

    def test_invalid_json_body(self):
        # 测试无效的 JSON 数据
        response = self.client.post(reverse('post-submission-url'), data="Invalid JSON", content_type='application/json')
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json().get('error'), 'Invalid JSON body')

#问卷展示界面：

class display_answer_normal_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Submitted')

        # 创建问题和答案
        self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
        self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)
        
        self.choice_option = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
        self.choice_answer = ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option)
        self.blank_answer = BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Test Answer")
        self.rating_answer = RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

        # 设置URL
        self.url = reverse('display-answer-normal', kwargs={'username': self.user.username, 'questionnaireId': self.survey.SurveyID, 'submissionId': self.submission.SubmissionID})

    def test_display_answer_normal_success(self):
        # 测试成功展示问卷答案
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertIn('Choice Question', response.json()['questionList'][0]['question'])
        self.assertIn('Blank Question', response.json()['questionList'][1]['question'])
        self.assertIn('Rating Question', response.json()['questionList'][2]['question'])

    def test_user_not_found(self):
        # 测试用户不存在的情况
        url = reverse('display-answer-normal', kwargs={'username': 'unknownuser', 'questionnaireId': self.survey.SurveyID, 'submissionId': self.submission.SubmissionID})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)
        self.assertIn('User not found', response.content.decode())

class display_answer_test_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, SubmissionID=1, Status='Submitted', Score=80)

        # 创建问题和答案
        self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
        self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)
        
        self.choice_option = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
        self.choice_answer = ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option)
        self.blank_answer = BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Test Answer")
        self.rating_answer = RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

        # 设置URL
        self.url = reverse('display-answer-test', kwargs={'username': self.user.username, 'questionnaireId': self.survey.SurveyID, 'submissionId': self.submission.SubmissionID})

    def test_display_answer_test_success(self):
        # 测试成功展示考试问卷答案
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        response_data = response.json()

        self.assertIn('Choice Question', response_data['questionList'][0]['question'])
        self.assertIn('Blank Question', response_data['questionList'][1]['question'])
        self.assertIn('Rating Question', response_data['questionList'][2]['question'])
        self.assertEqual(response_data['score'], 80)

    def test_survey_not_found(self):
        # 测试问卷不存在的情况
        url = reverse('display-answer-test', kwargs={'username': self.user.username, 'questionnaireId': 9999, 'submissionId': self.submission.SubmissionID})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)
        self.assertIn('Questionnaire not found', response.content.decode())

#数据分析:

class cross_analysis_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123', zhibi=100)
        self.survey = Survey.objects.create(Owner=self.user, Title="Test Survey", Category=1, TimeLimit=30, Is_released=True)

        # 创建问题
        self.question1 = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question 1", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.question2 = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question 2", Score=5, IsRequired=True, QuestionNumber=2, Category=1)

        # 创建选项
        self.option1_q1 = ChoiceOption.objects.create(Question=self.question1, Text="Option 1.1", IsCorrect=True, OptionNumber=1)
        self.option2_q1 = ChoiceOption.objects.create(Question=self.question1, Text="Option 1.2", IsCorrect=False, OptionNumber=2)

        self.option1_q2 = ChoiceOption.objects.create(Question=self.question2, Text="Option 2.1", IsCorrect=True, OptionNumber=1)
        self.option2_q2 = ChoiceOption.objects.create(Question=self.question2, Text="Option 2.2", IsCorrect=False, OptionNumber=2)

        # 创建提交记录和答案
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Submitted', Score=80)
        ChoiceAnswer.objects.create(Question=self.question1, Submission=self.submission, ChoiceOptions=self.option1_q1)
        ChoiceAnswer.objects.create(Question=self.question2, Submission=self.submission, ChoiceOptions=self.option1_q2)

        # 设置URL
        self.url = reverse('cross-analysis-url', kwargs={'QuestionID1': self.question1.QuestionID, 'QuestionID2': self.question2.QuestionID})

    def test_cross_analysis_success(self):
        # 测试交叉分析的成功情况
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        response_data = response.json()

        self.assertIn('crossCount', response_data)
        self.assertIn('crossText', response_data)
        self.assertEqual(len(response_data['crossCount']), 4)
        self.assertEqual(len(response_data['crossText']), 4)

    def test_questions_from_different_surveys(self):
        # 测试两个问题来自不同问卷的情况
        survey2 = Survey.objects.create(Owner=self.user, Title="Another Survey", Category=1, TimeLimit=30, Is_released=True)
        question3 = ChoiceQuestion.objects.create(Survey=survey2, Text="Choice Question 3", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        url = reverse('cross-analysis-url', kwargs={'QuestionID1': self.question1.QuestionID, 'QuestionID2': question3.QuestionID})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)
        self.assertIn('Two questions are not from the same questionnaire.', response.content.decode())

class survey_statistics_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123')
        self.survey = Survey.objects.create(Owner=self.user, Title="Survey Statistics", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)

        # 创建问题
        self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
        self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)

        # 创建选项
        self.choice_option_1 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
        self.choice_option_2 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 2", IsCorrect=False, OptionNumber=2)

        # 创建提交记录和答案
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Submitted', Score=80)
        ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option_1)
        BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Blank Answer")
        RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

        # 设置URL
        self.url = reverse('survey-statistics-url', kwargs={'surveyID': self.survey.SurveyID})

    def test_survey_statistics_success(self):
        # 测试获取问卷统计数据的成功情况
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        response_data = response.json()

        self.assertIn('title', response_data)
        self.assertEqual(response_data['title'], self.survey.Title)
        self.assertIn('questionList', response_data)
        self.assertEqual(len(response_data['questionList']), 3)

        # 检查选择题统计
        choice_question_data = next(item for item in response_data['questionList'] if item['type'] == 1)
        self.assertIn('Content', choice_question_data)
        self.assertEqual(choice_question_data['Content'], self.choice_question.Text)
        self.assertEqual(len(choice_question_data['Text']), 2)
        self.assertEqual(len(choice_question_data['Count']), 2)

        # 检查填空题统计
        blank_question_data = next(item for item in response_data['questionList'] if item['type'] == 3)
        self.assertIn('Content', blank_question_data)
        self.assertEqual(blank_question_data['Content'], self.blank_question.Text)
        self.assertEqual(blank_question_data['Text'][0], "Blank Answer")

        # 检查评分题统计
        rating_question_data = next(item for item in response_data['questionList'] if item['type'] == 4)
        self.assertIn('Content', rating_question_data)
        self.assertEqual(rating_question_data['Content'], self.rating_question.Text)
        self.assertEqual(rating_question_data['Text'][0], 4)

    def test_survey_not_found(self):
        # 测试问卷不存在的情况
        url = reverse('survey-statistics-url', kwargs={'surveyID': 999})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 404)
        self.assertIn('Survey not found', response.content.decode())

class download_submissions_test(TestCase):
    def setUp(self):
        # 设置测试环境
        self.client = APIClient()
        self.user = User.objects.create(username='testuser', email='user@example.com', password='password123')
        self.survey = Survey.objects.create(Owner=self.user, Title="Survey Statistics", Category=1, TimeLimit=30, SurveyID=1, Is_released=True)

        # 创建问题
        self.choice_question = ChoiceQuestion.objects.create(Survey=self.survey, Text="Choice Question", Score=5, IsRequired=True, QuestionNumber=1, Category=1)
        self.blank_question = BlankQuestion.objects.create(Survey=self.survey, Text="Blank Question", Score=10, IsRequired=True, QuestionNumber=2, Category=3)
        self.rating_question = RatingQuestion.objects.create(Survey=self.survey, Text="Rating Question", Score=5, IsRequired=True, QuestionNumber=3, Category=4)

        # 创建选项
        self.choice_option_1 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 1", IsCorrect=True, OptionNumber=1)
        self.choice_option_2 = ChoiceOption.objects.create(Question=self.choice_question, Text="Option 2", IsCorrect=False, OptionNumber=2)

        # 创建提交记录和答案
        self.submission = Submission.objects.create(Survey=self.survey, Respondent=self.user, Status='Submitted', Interval=0, Score=80)
        ChoiceAnswer.objects.create(Question=self.choice_question, Submission=self.submission, ChoiceOptions=self.choice_option_1)
        BlankAnswer.objects.create(Question=self.blank_question, Submission=self.submission, Content="Blank Answer")
        RatingAnswer.objects.create(Question=self.rating_question, Submission=self.submission, Rate=4)

        # 设置URL
        self.url = reverse('download_submissions-url', kwargs={'surveyID': self.survey.SurveyID})

    def test_download_submissions_success(self):
        # 测试成功下载提交记录
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.ms-excel')

        # 读取返回的Excel文件
        output = BytesIO(response.content)
        workbook = openpyxl.load_workbook(output)
        worksheet = workbook.active

        # 检查Excel中的内容
        self.assertEqual(worksheet.cell(1, 1).value, '填写者')
        self.assertEqual(worksheet.cell(1, 2).value, '提交答卷时间')
        self.assertEqual(worksheet.cell(1, 3).value, '所用时间')
        self.assertEqual(worksheet.cell(2, 1).value, self.user.username)
        self.assertEqual(worksheet.cell(2, 3).value, '--') # 对于非考试问卷，其所用时间为'--'
        self.assertEqual(worksheet.cell(2, 4).value, '1')  # 选项号1
        self.assertEqual(worksheet.cell(2, 5).value, 'Blank Answer')
        self.assertEqual(worksheet.cell(2, 6).value, 4)

    def test_no_submissions(self):
        # 测试无提交记录的情况
        Submission.objects.all().delete()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 404)
        self.assertIn('No submission records available.', response.content.decode())
